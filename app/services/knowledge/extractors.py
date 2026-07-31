"""Document text extractors for supported knowledge formats."""
from __future__ import annotations

import csv
import io
import json
import re
from html.parser import HTMLParser
from typing import Optional, Tuple
from urllib.parse import urlparse
from xml.etree import ElementTree as ET

import requests


class _HTMLTextExtractor(HTMLParser):
    def __init__(self):
        super().__init__()
        self._parts = []
        self._skip = False

    def handle_starttag(self, tag, attrs):
        if tag in ('script', 'style', 'noscript'):
            self._skip = True

    def handle_endtag(self, tag):
        if tag in ('script', 'style', 'noscript'):
            self._skip = False
        if tag in ('p', 'div', 'br', 'li', 'h1', 'h2', 'h3', 'tr'):
            self._parts.append('\n')

    def handle_data(self, data):
        if not self._skip:
            self._parts.append(data)

    def get_text(self) -> str:
        return re.sub(r'\n{3,}', '\n\n', ''.join(self._parts)).strip()


def clean_text(text: str) -> str:
    if not text:
        return ''
    text = text.replace('\r\n', '\n').replace('\r', '\n')
    text = re.sub(r'[ \t]+\n', '\n', text)
    text = re.sub(r'\n{3,}', '\n\n', text)
    text = re.sub(r'[ \t]{2,}', ' ', text)
    return text.strip()


def detect_doc_type(filename: str = '', mime: str = '', source_type: str = 'upload') -> str:
    name = (filename or '').lower()
    if source_type == 'url':
        return 'url'
    if source_type == 'sitemap':
        return 'sitemap'
    if source_type == 'manual' or source_type == 'note':
        return 'note'
    mapping = {
        '.pdf': 'pdf', '.docx': 'docx', '.txt': 'txt', '.csv': 'csv',
        '.xlsx': 'xlsx', '.xls': 'xlsx', '.md': 'markdown', '.markdown': 'markdown',
        '.html': 'html', '.htm': 'html', '.json': 'json', '.xml': 'xml',
    }
    for ext, dtype in mapping.items():
        if name.endswith(ext):
            return dtype
    mime = (mime or '').lower()
    if 'pdf' in mime:
        return 'pdf'
    if 'word' in mime or 'docx' in mime:
        return 'docx'
    if 'html' in mime:
        return 'html'
    if 'json' in mime:
        return 'json'
    if 'csv' in mime:
        return 'csv'
    if 'sheet' in mime or 'excel' in mime:
        return 'xlsx'
    return 'txt'


def extract_from_bytes(data: bytes, doc_type: str, filename: str = '') -> str:
    if doc_type == 'pdf':
        return _extract_pdf(data)
    if doc_type == 'docx':
        return _extract_docx(data)
    if doc_type == 'xlsx':
        return _extract_xlsx(data)
    if doc_type == 'csv':
        return _extract_csv(data)
    if doc_type == 'json':
        return _extract_json(data)
    if doc_type in ('html',):
        return _extract_html(data.decode('utf-8', errors='ignore'))
    if doc_type in ('markdown', 'txt', 'note', 'brand', 'marketing', 'research', 'case_study', 'playbook', 'xml'):
        return clean_text(data.decode('utf-8', errors='ignore'))
    return clean_text(data.decode('utf-8', errors='ignore'))


def extract_from_url(url: str, timeout: int = 20) -> Tuple[str, str]:
    """Returns (text, detected_type)."""
    resp = requests.get(url, timeout=timeout, headers={'User-Agent': 'OplyraKnowledgeBot/1.0'})
    resp.raise_for_status()
    ctype = (resp.headers.get('Content-Type') or '').lower()
    if 'xml' in ctype or url.rstrip('/').endswith('.xml') or 'sitemap' in url.lower():
        return extract_sitemap_xml(resp.text), 'sitemap'
    if 'html' in ctype or urlparse(url).path.endswith(('.html', '.htm', '')):
        return _extract_html(resp.text), 'html'
    if 'json' in ctype:
        return _extract_json(resp.content), 'json'
    return clean_text(resp.text), 'url'


def extract_sitemap_xml(xml_text: str) -> str:
    urls = []
    try:
        root = ET.fromstring(xml_text)
        for el in root.iter():
            if el.tag.endswith('loc') and el.text:
                urls.append(el.text.strip())
    except ET.ParseError:
        urls = re.findall(r'<loc>(.*?)</loc>', xml_text)
    lines = ['Sitemap URLs:'] + [f'- {u}' for u in urls[:500]]
    return '\n'.join(lines)


def _extract_pdf(data: bytes) -> str:
    try:
        from pypdf import PdfReader
        reader = PdfReader(io.BytesIO(data))
        parts = []
        for page in reader.pages:
            parts.append(page.extract_text() or '')
        return clean_text('\n\n'.join(parts))
    except ImportError:
        # Minimal fallback: extract readable strings
        text = data.decode('latin-1', errors='ignore')
        strings = re.findall(r'[\x20-\x7E]{4,}', text)
        return clean_text('\n'.join(strings[:2000]))


def _extract_docx(data: bytes) -> str:
    try:
        from docx import Document
        doc = Document(io.BytesIO(data))
        return clean_text('\n'.join(p.text for p in doc.paragraphs if p.text))
    except Exception:
        return clean_text(data.decode('utf-8', errors='ignore'))


def _extract_xlsx(data: bytes) -> str:
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        parts = []
        for sheet in wb.worksheets:
            parts.append(f'# Sheet: {sheet.title}')
            for row in sheet.iter_rows(values_only=True):
                cells = [str(c) for c in row if c is not None]
                if cells:
                    parts.append('\t'.join(cells))
        return clean_text('\n'.join(parts))
    except ImportError:
        return clean_text('Excel file uploaded (openpyxl not installed — install for full extraction).')
    except Exception as exc:
        return clean_text(f'Excel extraction error: {exc}')


def _extract_csv(data: bytes) -> str:
    text = data.decode('utf-8', errors='ignore')
    reader = csv.reader(io.StringIO(text))
    rows = []
    for i, row in enumerate(reader):
        if i > 5000:
            rows.append('… truncated …')
            break
        rows.append(', '.join(row))
    return clean_text('\n'.join(rows))


def _extract_json(data: bytes) -> str:
    try:
        obj = json.loads(data.decode('utf-8', errors='ignore'))
        return clean_text(json.dumps(obj, indent=2, ensure_ascii=False))
    except Exception:
        return clean_text(data.decode('utf-8', errors='ignore'))


def _extract_html(html: str) -> str:
    parser = _HTMLTextExtractor()
    try:
        parser.feed(html)
        return clean_text(parser.get_text())
    except Exception:
        return clean_text(re.sub(r'<[^>]+>', ' ', html))
